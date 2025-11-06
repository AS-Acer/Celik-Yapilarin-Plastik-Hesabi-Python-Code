# -*- coding: utf-8 -*-
"""
Section properties & capacities (elastic/plastic) with stylish console output
+ robust Desktop save (handles OneDrive) + CSV and Excel (.xlsx) export.

Covers three section types:
(1) Built-up I (asymmetric)
(2) CHS + 2×UPE300 (left-right)  -> gap_back = 28.9 (updated)
(3) CHS + 2×UPE300 (top-bottom)  -> y_c = 190.4  (updated)

Units:
- Geometry in mm, stresses in MPa (N/mm^2).
- Moments printed in kN·m; Ix,Iy in mm^4; We,Wp in mm^3.
"""
from dataclasses import dataclass, field
from typing import Dict, Any, List, Tuple
import math
import csv
import os
from pathlib import Path

# ==========================
# Try Excel writer
# ==========================
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side, NamedStyle
    OPENPYXL_OK = True
except Exception:
    OPENPYXL_OK = False

# ==========================
# Config
# ==========================
VERBOSE = True
SAVE_CSV = True
SAVE_XLSX = True

APP_NAME = "sections_results"

def find_desktop_path() -> Path:
    """
    Find Desktop path robustly:
    1) %USERPROFILE%\Desktop
    2) %OneDrive%\Desktop
    3) current working directory as fallback
    """
    user = Path.home()
    p1 = user / "Desktop"
    if p1.exists():
        return p1

    onedrive = os.environ.get("OneDrive")
    if onedrive:
        p2 = Path(onedrive) / "Desktop"
        if p2.exists():
            return p2

    # Fallback
    return Path.cwd()

DESKTOP = find_desktop_path()
CSV_PATH = DESKTOP / f"{APP_NAME}.csv"
XLSX_PATH = DESKTOP / f"{APP_NAME}.xlsx"

# ==========================
# Pretty printing helpers
# ==========================
def nmm_to_knm(M_nmm: float) -> float:
    return M_nmm / 1e6

def line(w: int = 64, ch: str = "─") -> str:
    return ch * w

def box(title: str, width: int = 64) -> str:
    t = f" {title} "
    pad = max(0, width - len(t) - 2)
    return f"┌{t}{'─'*pad}┐"

def box_close(width: int = 64) -> str:
    return f"└{line(width-2)}┘"

def kv(k: str, v: Any, unit: str = "", k_w: int = 20, v_w: int = 18) -> str:
    sval = f"{v}"
    if isinstance(v, float):
        sval = f"{v:.6g}"
    unit_str = f" {unit}" if unit else ""
    return f"{k:<{k_w}} : {sval:>{v_w}}{unit_str}"

def as_table(rows: List[Tuple[str, Any, str]], title: str = "", width: int = 64) -> str:
    out = []
    if title:
        out.append(box(title, width))
    for k, v, unit in rows:
        out.append(kv(k, v, unit))
    if title:
        out.append(box_close(width))
    return "\n".join(out)

def section_header(name: str) -> None:
    print("\n" + box(name))
    print("│ " + "Computation & Results".ljust(60) + "│")
    print(box_close())

def explain(msg: str) -> None:
    if VERBOSE:
        for line_ in msg.strip().splitlines():
            print(f"  • {line_.strip()}")

# ==========================
# Material
# ==========================
@dataclass
class Material:
    """Material model (only yield stress used)."""
    fy: float = 355.0  # MPa = N/mm^2

def _mat_default() -> Material:
    return Material(fy=355.0)

# ==========================
# UPE300 Catalog
# ==========================
@dataclass
class UPE300Catalog:
    """Minimal catalog for UPE300 with geometric & stiffness props."""
    A: float = 5660.0          # mm^2
    Ix: float = 3.24e7         # mm^4  (strong axis)  3240 cm^4 -> 3.24e7 mm^4
    Iy: float = 1.52e6         # mm^4  (weak axis)
    h: float = 300.0           # mm
    b: float = 100.0           # mm
    tf: float = 15.0           # mm
    tw: float = 9.5            # mm
    hi: float = 270.0          # mm (~ h - 2*tf)
    e_back_to_cg: float = 27.5 # mm (back face -> CG, for LR case)

    @property
    def We_x(self) -> float:
        return self.Ix / (self.h / 2.0)

    @property
    def Wp_x_est(self) -> float:
        return 1.12 * self.We_x   # channel shape factor ~1.10–1.15

    @property
    def Wp_y_exact(self) -> float:
        return 2.0 * (self.tf * self.b**2 / 4.0) + (self.hi * self.tw**2 / 4.0)

def _upe_default() -> UPE300Catalog:
    return UPE300Catalog()

# ==========================
# Section 1: Built-up I
# ==========================
@dataclass
class BuiltUpI:
    b_top: float
    t_top: float
    b_bot: float
    t_bot: float
    tw: float
    H: float
    mat: Material = field(default_factory=_mat_default)

    def compute(self) -> Dict[str, float]:
        if VERBOSE:
            explain("Section 1: Built-up I — compute areas and centroid (needed for Ix).")
        A_top = self.b_top * self.t_top
        A_web = self.tw * (self.H - self.t_top - self.t_bot)
        A_bot = self.b_bot * self.t_bot
        A = A_top + A_web + A_bot

        y_top = self.t_top / 2.0
        y_web = self.t_top + (self.H - self.t_top - self.t_bot) / 2.0
        y_bot = self.H - self.t_bot / 2.0

        ybar = (A_top*y_top + A_web*y_web + A_bot*y_bot) / A
        c_top = ybar
        c_bot = self.H - ybar

        if VERBOSE:
            explain("Parallel-axis (Steiner) for Ix about centroidal x-axis.")

        Ix_top0 = self.b_top * self.t_top**3 / 12.0
        Ix_web0 = self.tw * (self.H - self.t_top - self.t_bot)**3 / 12.0
        Ix_bot0 = self.b_bot * self.t_bot**3 / 12.0
        d_top = abs(y_top - ybar)
        d_web = abs(y_web - ybar)
        d_bot = abs(y_bot - ybar)
        Ix = (Ix_top0 + A_top*d_top**2
              + Ix_web0 + A_web*d_web**2
              + Ix_bot0 + A_bot*d_bot**2)

        Iy_top = self.t_top * self.b_top**3 / 12.0
        Iy_web = (self.H - self.t_top - self.t_bot) * self.tw**3 / 12.0
        Iy_bot = self.t_bot * self.b_bot**3 / 12.0
        Iy = Iy_top + Iy_web + Iy_bot

        We_x = min(Ix / c_top, Ix / c_bot)
        We_y = Iy / (max(self.b_top/2.0, self.b_bot/2.0))

        if VERBOSE:
            explain("Plastic neutral axis (x): balance compression and tension areas.")
        A_half = A / 2.0
        if A_top >= A_half:
            t = A_half / self.b_top
            y_PNA = t
            Qc = self.b_top * t * (t/2.0)
        else:
            rem = A_half - A_top
            t = rem / self.tw
            y_PNA = self.t_top + t
            Qc_top = A_top * (y_PNA - y_top)
            Qc_web = (self.tw * t) * (t/2.0)
            Qc = Qc_top + Qc_web

        web_lower_h = (self.H - self.t_top - self.t_bot) - (y_PNA - self.t_top)
        Qt_web = (self.tw * web_lower_h) * (web_lower_h/2.0)
        dist_to_bot = (self.H - self.t_bot/2.0) - y_PNA
        Qt_bot = A_bot * dist_to_bot
        Wp_x = Qc + Qt_web + Qt_bot

        Wp_y = (self.t_top * self.b_top**2 / 4.0
                + self.t_bot * self.b_bot**2 / 4.0
                + (self.H - self.t_top - self.t_bot) * self.tw**2 / 4.0)

        fy = self.mat.fy
        return {
            "A_mm2": A,
            "Ix_mm4": Ix, "Iy_mm4": Iy,
            "We_x_mm3": We_x, "We_y_mm3": We_y,
            "Wp_x_mm3": Wp_x, "Wp_y_mm3": Wp_y,
            "Me_x_kNm": nmm_to_knm(fy * We_x),
            "Mp_x_kNm": nmm_to_knm(fy * Wp_x),
            "shape_x": Wp_x/We_x,
            "Me_y_kNm": nmm_to_knm(fy * We_y),
            "Mp_y_kNm": nmm_to_knm(fy * Wp_y),
            "shape_y": Wp_y/We_y,
        }

# ==========================
# CHS helper (tube)
# ==========================
@dataclass
class CHS:
    Do: float  # mm
    t: float   # mm

    @property
    def Ro(self) -> float:
        return self.Do / 2.0

    @property
    def Ri(self) -> float:
        return self.Do / 2.0 - self.t

    @property
    def A(self) -> float:
        return math.pi * (self.Ro**2 - self.Ri**2)

    @property
    def I(self) -> float:
        return (math.pi/4.0) * (self.Ro**4 - self.Ri**4)

    @property
    def Wp(self) -> float:
        return (4.0/3.0) * (self.Ro**3 - self.Ri**3)

# ==========================
# Section 2: CHS + 2×UPE (L-R)
# ==========================
@dataclass
class CHS_UPE_LR:
    tube: CHS
    upe: UPE300Catalog = field(default_factory=_upe_default)
    gap_back: float = 28.9
    mat: Material = field(default_factory=_mat_default)
    Wp_x_upe_override: float = None

    def compute(self) -> Dict[str, float]:
        Ro = self.tube.Ro
        x_c = Ro + self.gap_back + self.upe.e_back_to_cg

        if VERBOSE:
            explain("Section 2: CG is at tube center by LR symmetry (x & y).")

        A_total = self.tube.A + 2.0 * self.upe.A

        Ix = self.tube.I + 2.0 * self.upe.Ix
        Iy = self.tube.I + 2.0 * (self.upe.Iy + self.upe.A * x_c**2)

        c_x = max(Ro, self.upe.h/2.0)
        x_outer_upe = max(Ro + self.gap_back, x_c + self.upe.b/2.0)
        c_y = max(Ro, x_outer_upe)

        We_x = Ix / c_x
        We_y = Iy / c_y

        Wp_tube = self.tube.Wp
        Wp_x_upe = self.Wp_x_upe_override if self.Wp_x_upe_override is not None else self.upe.Wp_x_est
        Wp_x = Wp_tube + 2.0 * Wp_x_upe
        Wp_y = Wp_tube + 2.0 * (self.upe.A * x_c)

        fy = self.mat.fy
        return {
            "section": "CHS+2×UPE L-R",
            "x_c_mm": x_c,
            "A_mm2": A_total,
            "Ix_mm4": Ix, "Iy_mm4": Iy,
            "We_x_mm3": We_x, "We_y_mm3": We_y,
            "Wp_x_mm3": Wp_x, "Wp_y_mm3": Wp_y,
            "Me_x_kNm": nmm_to_knm(fy * We_x),
            "Mp_x_kNm": nmm_to_knm(fy * Wp_x),
            "shape_x": Wp_x/We_x,
            "Me_y_kNm": nmm_to_knm(fy * We_y),
            "Mp_y_kNm": nmm_to_knm(fy * Wp_y),
            "shape_y": Wp_y/We_y,
        }

# ==========================
# Section 3: CHS + 2×UPE (T-B)
# ==========================
@dataclass
class CHS_UPE_TB:
    tube: CHS
    upe: UPE300Catalog = field(default_factory=_upe_default)
    y_c: float = 190.4
    mat: Material = field(default_factory=_mat_default)

    def compute(self) -> Dict[str, float]:
        Ro = self.tube.Ro
        if VERBOSE:
            explain("Section 3: by TB symmetry, global CG is at tube center (0,0).")
            explain("Ix gets Steiner (A*y_c^2); Iy does not (no x-offset).")

        Ix = self.tube.I + 2.0 * (self.upe.Ix + self.upe.A * self.y_c**2)
        Iy = self.tube.I + 2.0 * self.upe.Iy

        c_x = max(Ro, self.y_c + self.upe.h/2.0)
        c_y = max(Ro, self.upe.b/2.0)

        We_x = Ix / c_x
        We_y = Iy / c_y

        Wp_tube = self.tube.Wp
        Wp_x = Wp_tube + 2.0 * (self.upe.A * self.y_c)
        Wp_y = Wp_tube + 2.0 * self.upe.Wp_y_exact

        fy = self.mat.fy
        return {
            "section": "CHS+2×UPE T-B",
            "y_c_mm": self.y_c,
            "A_mm2": self.tube.A + 2.0 * self.upe.A,
            "Ix_mm4": Ix, "Iy_mm4": Iy,
            "We_x_mm3": We_x, "We_y_mm3": We_y,
            "Wp_x_mm3": Wp_x, "Wp_y_mm3": Wp_y,
            "Me_x_kNm": nmm_to_knm(fy * We_x),
            "Mp_x_kNm": nmm_to_knm(fy * Wp_x),
            "shape_x": Wp_x/We_x,
            "Me_y_kNm": nmm_to_knm(fy * We_y),
            "Mp_y_kNm": nmm_to_knm(fy * Wp_y),
            "shape_y": Wp_y/We_y,
        }

# ==========================
# CSV util (UNION of keys)
# ==========================
def save_csv(rows: List[Dict[str, Any]], path: Path) -> None:
    if not rows:
        return
    # Union of keys
    all_keys = set()
    for r in rows:
        all_keys.update(r.keys())

    preferred = [
        "section", "A_mm2",
        "Ix_mm4", "Iy_mm4",
        "We_x_mm3", "Wp_x_mm3", "Me_x_kNm", "Mp_x_kNm", "shape_x",
        "We_y_mm3", "Wp_y_mm3", "Me_y_kNm", "Mp_y_kNm", "shape_y",
        "x_c_mm", "y_c_mm"
    ]
    keys = [k for k in preferred if k in all_keys] + [k for k in sorted(all_keys) if k not in preferred]

    path.parent.mkdir(parents=True, exist_ok=True)
    with open(path, "w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=keys)
        w.writeheader()
        for r in rows:
            row_out = {k: r.get(k, "") for k in keys}
            w.writerow(row_out)

# ==========================
# Excel util (openpyxl)
# ==========================
def save_xlsx(rows: List[Dict[str, Any]], path: Path) -> None:
    if not rows:
        return
    if not OPENPYXL_OK:
        print("(!) openpyxl not installed. Install with: pip install openpyxl")
        return

    # Union + stable order
    all_keys = set()
    for r in rows:
        all_keys.update(r.keys())
    preferred = [
        "section", "A_mm2",
        "Ix_mm4", "Iy_mm4",
        "We_x_mm3", "Wp_x_mm3", "Me_x_kNm", "Mp_x_kNm", "shape_x",
        "We_y_mm3", "Wp_y_mm3", "Me_y_kNm", "Mp_y_kNm", "shape_y",
        "x_c_mm", "y_c_mm"
    ]
    headers = [k for k in preferred if k in all_keys] + [k for k in sorted(all_keys) if k not in preferred]

    wb = Workbook()
    ws = wb.active
    ws.title = "Summary"

    # Styles
    bold = Font(bold=True)
    center = Alignment(horizontal="center", vertical="center")
    hdr_fill = PatternFill("solid", fgColor="D9E1F2")
    thin = Side(style="thin", color="999999")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Header
    ws.append(headers)
    for c in ws[1]:
        c.font = bold
        c.alignment = center
        c.fill = hdr_fill
        c.border = border

    # Data
    for r in rows:
        ws.append([r.get(h, "") for h in headers])

    # Column widths (simple autosize)
    for col in ws.columns:
        length = max(len(str(c.value)) if c.value is not None else 0 for c in col)
        ws.column_dimensions[col[0].column_letter].width = min(max(12, length + 2), 28)

    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)

# ==========================
# Reporting (console)
# ==========================
def report(title: str, results: Dict[str, Any]) -> None:
    rows = [
        ("Area", results.get("A_mm2", "—"), "mm²"),
        ("Ix", results["Ix_mm4"], "mm⁴"),
        ("Iy", results["Iy_mm4"], "mm⁴"),
        ("We_x", results["We_x_mm3"], "mm³"),
        ("Wp_x", results["Wp_x_mm3"], "mm³"),
        ("Me_x", results["Me_x_kNm"], "kN·m"),
        ("Mp_x", results["Mp_x_kNm"], "kN·m"),
        ("shape_x", results["shape_x"], "—"),
        ("We_y", results["We_y_mm3"], "mm³"),
        ("Wp_y", results["Wp_y_mm3"], "mm³"),
        ("Me_y", results["Me_y_kNm"], "kN·m"),
        ("Mp_y", results["Mp_y_kNm"], "kN·m"),
        ("shape_y", results["shape_y"], "—"),
    ]
    section_header(title)
    print(as_table(rows, title="Section Summary"))

# ==========================
# Demo / main
# ==========================
if __name__ == "__main__":
    print(box(f"Save folder → {DESKTOP}"))
    print(box_close())

    # SECTION 1
    s1 = BuiltUpI(b_top=300, t_top=20, b_bot=100, t_bot=20, tw=10, H=800)
    r1 = s1.compute()
    report("Section 1 - Built-up I", r1)
    r1["section"] = "Built-up I"

    # SECTION 2 (LR) with updated gap_back=28.9
    tube = CHS(Do=323.0, t=12.0)
    s2 = CHS_UPE_LR(tube=tube)
    r2 = s2.compute()
    if VERBOSE:
        explain("For LR: c_x = max(Ro, h/2);  c_y = max(Ro, x_outer_UPE).")
        explain("Steiner affects Iy via A*x_c^2; Ix only sums local Ix.")
    report("Section 2 - CHS + 2×UPE300 (L-R)", r2)

    # SECTION 3 (TB) with updated y_c=190.4
    s3 = CHS_UPE_TB(tube=tube)
    r3 = s3.compute()
    if VERBOSE:
        explain("For TB: c_x = y_c + h/2 (top UPE governs); c_y = max(Ro, b/2).")
        explain("Steiner affects Ix via A*y_c^2; Iy only sums local Iy.")
    report("Section 3 - CHS + 2×UPE300 (T-B)", r3)

    # Collect & save
    rows: List[Dict[str, Any]] = [r1, r2, r3]

    if SAVE_CSV:
        save_csv(rows, CSV_PATH)
        print("\n" + box(f"Saved CSV → {CSV_PATH}"))
        print(box_close())

    if SAVE_XLSX:
        if OPENPYXL_OK:
            save_xlsx(rows, XLSX_PATH)
            print("\n" + box(f"Saved Excel → {XLSX_PATH}"))
            print(box_close())
        else:
            print("\n" + box("Excel save skipped (install: pip install openpyxl)"))
            print(box_close())
